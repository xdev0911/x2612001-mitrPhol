"""
SKU Router
==========
SKU (recipe) management, steps, actions, destinations, and phases.
"""

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from typing import List
from datetime import datetime
import os
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

import crud
import models
import schemas
from database import get_db

logger = logging.getLogger(__name__)
router = APIRouter(tags=["SKUs"])


# =============================================================================
# SKU MASTER ENDPOINTS
# =============================================================================

@router.get("/skus/", response_model=List[schemas.Sku])
def get_skus(skip: int = 0, limit: int = 100, sku_id: str = None, db: Session = Depends(get_db)):
    """Get all SKUs with pagination and optional filter."""
    skip = max(0, skip)
    limit = min(max(1, limit), 1000)
    
    if sku_id:
        sku = crud.get_sku_by_sku_id(db, sku_id=sku_id)
        return [sku] if sku else []
    
    return crud.get_skus(db, skip=skip, limit=limit)


@router.get("/skus/{sku_db_id}", response_model=schemas.Sku)
def get_sku(sku_db_id: int, db: Session = Depends(get_db)):
    """Get SKU by ID."""
    db_sku = db.query(models.Sku).filter(models.Sku.id == sku_db_id).first()
    if db_sku is None:
        raise HTTPException(status_code=404, detail="SKU not found")
    return db_sku


@router.post("/skus/", response_model=schemas.Sku)
def create_sku(sku: schemas.SkuCreate, db: Session = Depends(get_db)):
    """Create new SKU."""
    try:
        if crud.get_sku_by_sku_id(db, sku_id=sku.sku_id):
            raise HTTPException(status_code=400, detail="SKU ID already exists")
        return crud.create_sku(db=db, sku=sku)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except RuntimeError:
        raise HTTPException(status_code=500, detail="Database error")


@router.put("/skus/{sku_db_id}", response_model=schemas.Sku)
def update_sku(sku_db_id: int, sku: schemas.SkuCreate, db: Session = Depends(get_db)):
    """Update SKU."""
    try:
        db_sku = crud.update_sku(db, sku_db_id=sku_db_id, sku_update=sku)
        if db_sku is None:
            raise HTTPException(status_code=404, detail="SKU not found")
        return db_sku
    except RuntimeError:
        raise HTTPException(status_code=500, detail="Database error")


@router.delete("/skus/{sku_db_id}")
def delete_sku(sku_db_id: int, db: Session = Depends(get_db)):
    """Soft delete SKU by setting status to 'Deleted'."""
    try:
        db_sku = db.query(models.Sku).filter(models.Sku.id == sku_db_id).first()
        if db_sku is None:
            raise HTTPException(status_code=404, detail="SKU not found")
        
        db_sku.status = "Deleted"
        db.commit()
        return {"status": "success", "message": "SKU marked as deleted"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")


# =============================================================================
# SKU STEP ENDPOINTS
# =============================================================================

@router.get("/sku-steps/", response_model=List[schemas.SkuStep])
def get_sku_steps(sku_id: str = None, skip: int = 0, limit: int = 100, db: Session = Depends(get_db)):
    """Get all SKU steps, optionally filtered by sku_id."""
    query = db.query(models.SkuStep)
    if sku_id:
        query = query.filter(models.SkuStep.sku_id == sku_id)
    return query.offset(skip).limit(limit).all()


@router.post("/sku-steps/", response_model=schemas.SkuStep)
def create_sku_step(step: schemas.SkuStepCreate, db: Session = Depends(get_db)):
    """Create new SKU step."""
    try:
        db_step = models.SkuStep(**step.model_dump())
        db.add(db_step)
        db.commit()
        db.refresh(db_step)
        return db_step
    except Exception as e:
        logger.error(f"Error creating SKU step: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/sku-steps/{step_id}", response_model=schemas.SkuStep)
def get_sku_step(step_id: int, db: Session = Depends(get_db)):
    """Get SKU step by ID."""
    db_step = db.query(models.SkuStep).filter(models.SkuStep.id == step_id).first()
    if db_step is None:
        raise HTTPException(status_code=404, detail="Step not found")
    return db_step


@router.put("/sku-steps/{step_id}", response_model=schemas.SkuStep)
def update_sku_step(step_id: int, step: schemas.SkuStepCreate, db: Session = Depends(get_db)):
    """Update SKU step."""
    db_step = db.query(models.SkuStep).filter(models.SkuStep.id == step_id).first()
    if db_step is None:
        raise HTTPException(status_code=404, detail="Step not found")
    
    for key, value in step.dict().items():
        setattr(db_step, key, value)
    
    db.commit()
    db.refresh(db_step)
    return db_step


@router.delete("/sku-steps/{step_id}")
def delete_sku_step(step_id: int, db: Session = Depends(get_db)):
    """Delete SKU step."""
    db_step = db.query(models.SkuStep).filter(models.SkuStep.id == step_id).first()
    if db_step is None:
        raise HTTPException(status_code=404, detail="Step not found")
    
    db.delete(db_step)
    db.commit()
    return {"status": "success"}


# =============================================================================
# SKU ACTIONS, DESTINATIONS, PHASES
# =============================================================================

@router.get("/sku-actions/", response_model=List[schemas.SkuAction])
def get_sku_actions(skip: int = 0, limit: int = 100, db: Session = Depends(get_db)):
    return crud.get_sku_actions(db, skip=skip, limit=limit)


@router.post("/sku-actions/", response_model=schemas.SkuAction)
def create_sku_action(action: schemas.SkuActionCreate, db: Session = Depends(get_db)):
    try:
        return crud.create_sku_action(db, action=action)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.put("/sku-actions/{action_code}", response_model=schemas.SkuAction)
def update_sku_action(action_code: str, action: schemas.SkuActionCreate, db: Session = Depends(get_db)):
    result = crud.update_sku_action(db, action_code=action_code, action_update=action)
    if result is None:
        raise HTTPException(status_code=404, detail="Action not found")
    return result


@router.delete("/sku-actions/{action_code}")
def delete_sku_action(action_code: str, db: Session = Depends(get_db)):
    result = crud.delete_sku_action(db, action_code=action_code)
    if result is None:
        raise HTTPException(status_code=404, detail="Action not found")
    return {"status": "success"}


@router.get("/sku-destinations/", response_model=List[schemas.SkuDestination])
def get_sku_destinations(skip: int = 0, limit: int = 100, db: Session = Depends(get_db)):
    return crud.get_sku_destinations(db, skip=skip, limit=limit)


@router.post("/sku-destinations/", response_model=schemas.SkuDestination)
def create_sku_destination(dest: schemas.SkuDestinationCreate, db: Session = Depends(get_db)):
    try:
        return crud.create_sku_destination(db, dest=dest)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.put("/sku-destinations/{dest_id}", response_model=schemas.SkuDestination)
def update_sku_destination(dest_id: int, dest: schemas.SkuDestinationCreate, db: Session = Depends(get_db)):
    result = crud.update_sku_destination(db, dest_id=dest_id, dest_update=dest)
    if result is None:
        raise HTTPException(status_code=404, detail="Destination not found")
    return result


@router.delete("/sku-destinations/{dest_id}")
def delete_sku_destination(dest_id: int, db: Session = Depends(get_db)):
    result = crud.delete_sku_destination(db, dest_id=dest_id)
    if result is None:
        raise HTTPException(status_code=404, detail="Destination not found")
    return {"status": "success"}


@router.get("/sku-phases/", response_model=List[schemas.SkuPhase])
def get_sku_phases(skip: int = 0, limit: int = 100, db: Session = Depends(get_db)):
    return crud.get_sku_phases(db, skip=skip, limit=limit)


@router.post("/sku-phases/", response_model=schemas.SkuPhase)
def create_sku_phase(phase: schemas.SkuPhaseCreate, db: Session = Depends(get_db)):
    try:
        return crud.create_sku_phase(db, phase=phase)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.put("/sku-phases/{phase_id}", response_model=schemas.SkuPhase)
def update_sku_phase(phase_id: str, phase: schemas.SkuPhaseCreate, db: Session = Depends(get_db)):
    result = crud.update_sku_phase(db, phase_id=phase_id, phase_update=phase)
    if result is None:
        raise HTTPException(status_code=404, detail="Phase not found")
    return result


@router.delete("/sku-phases/{phase_id}")
def delete_sku_phase(phase_id: str, db: Session = Depends(get_db)):
    result = crud.delete_sku_phase(db, phase_id=phase_id)
    if result is None:
        raise HTTPException(status_code=404, detail="Phase not found")
    return {"status": "success"}


# =============================================================================
# SKU EXPORT
# =============================================================================

@router.get("/skus/export")
def export_skus_to_excel(sku_ids: str = None, db: Session = Depends(get_db)):
    """Export SKUs to Excel file."""
    try:
        query = db.query(models.Sku)
        if sku_ids:
            query = query.filter(models.Sku.sku_id.in_(sku_ids.split(",")))
        
        skus = query.all()
        wb = Workbook()
        ws = wb.active
        ws.title = "SKU Export"
        
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        header_font = Font(bold=True)
        
        headers = [
            "SKU ID", "SKU Name", "Status", "Created By", "Created At",
            "Updated By", "Updated At", "Step #", "Action Code", "Step Description",
            "Material", "Setup Steps", "Destination", "Setpoint", "Tol (+)",
            "Tol (-)", "Control Param", "Min Val", "Max Val", "Timer (s)",
            "Time Over Action", "QC Check"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        row_idx = 2
        for sku in skus:
            steps = sku.steps if sku.steps else [None]
            for step in steps:
                ws.cell(row=row_idx, column=1, value=sku.sku_id)
                ws.cell(row=row_idx, column=2, value=sku.sku_name)
                ws.cell(row=row_idx, column=3, value=sku.status)
                ws.cell(row=row_idx, column=4, value=sku.creat_by)
                ws.cell(row=row_idx, column=5, value=str(sku.created_at) if sku.created_at else "")
                ws.cell(row=row_idx, column=6, value=sku.update_by or "")
                ws.cell(row=row_idx, column=7, value=str(sku.updated_at) if sku.updated_at else "")
                
                if step:
                    step_num = f"{step.phase_number}.{step.sub_step}" if step.phase_number else str(step.sub_step)
                    ws.cell(row=row_idx, column=8, value=step_num)
                    ws.cell(row=row_idx, column=9, value=step.action_code)
                    ws.cell(row=row_idx, column=10, value=step.action_description or "")
                    ws.cell(row=row_idx, column=11, value=step.re_code or "")
                    ws.cell(row=row_idx, column=12, value=str(step.setup_step) if step.setup_step else "")
                    ws.cell(row=row_idx, column=13, value=step.destination or "")
                    ws.cell(row=row_idx, column=14, value=step.require)
                    ws.cell(row=row_idx, column=15, value=step.high_tol)
                    ws.cell(row=row_idx, column=16, value=step.low_tol)
                    ws.cell(row=row_idx, column=17, value=step.step_condition or "")
                    ws.cell(row=row_idx, column=18, value=step.temp_low)
                    ws.cell(row=row_idx, column=19, value=step.temp_high)
                    ws.cell(row=row_idx, column=20, value=step.step_time)
                    ws.cell(row=row_idx, column=21, value="")
                    
                    qc_checks = []
                    if step.qc_temp: qc_checks.append("QC Temp")
                    if step.record_steam_pressure: qc_checks.append("Steam Pressure")
                    if step.record_ctw: qc_checks.append("CTW")
                    if step.operation_brix_record: qc_checks.append("Brix")
                    if step.operation_ph_record: qc_checks.append("pH")
                    ws.cell(row=row_idx, column=22, value=", ".join(qc_checks))
                
                row_idx += 1
        
        export_dir = "/xApp/export"
        os.makedirs(export_dir, exist_ok=True)
        file_path = os.path.join(export_dir, "sku.xlsx")
        wb.save(file_path)
        
        return FileResponse(
            path=file_path,
            filename="sku.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except Exception as e:
        logger.error(f"SKU export failed: {e}")
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")
